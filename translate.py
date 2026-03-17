#!/usr/bin/env python3
"""
Paperturn Batch Translation Script
Scrapes English pages, translates via Claude API (translate + copywrite), outputs .xlsx files.
"""

import argparse
import csv
import os
import re
import sys
from pathlib import Path
from urllib.parse import urljoin, urlparse

import anthropic
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
BASE_URL = "https://www.paperturn.com"
MODEL = "claude-sonnet-4-6"
DEFAULT_OUTPUT = Path.home() / "Desktop" / "Translation Project" / "output"

TRANSLATOR_SYSTEM = (
    "You are an expert translator specializing in {lang}. "
    "Translate the following marketing content from English to {lang}. "
    "Preserve the exact meaning, tone, and intent. Do not add or remove information. "
    "CRITICAL: Keep the labels/tags in English exactly as they appear (e.g. 'H1:', 'P:', "
    "'CTA Button:', 'Page Title:', 'Meta Description:', 'URL Slug:'). "
    "Only translate the text AFTER the colon. "
    "Return the translation in the exact same structured format, one item per line."
)

COPYWRITER_SYSTEM = (
    "You are one of the world's leading copywriters and a native {lang} speaker "
    "who creates exceptional marketing copy. You have received a translation from English. "
    "Your job is to polish this into compelling, natural-sounding {lang} marketing copy "
    "that reads as if it were originally written in {lang}. "
    "Preserve the meaning and structure. "
    "CRITICAL: Keep the labels/tags in English exactly as they appear (e.g. 'H1:', 'P:', "
    "'CTA Button:', 'Page Title:', 'Meta Description:', 'URL Slug:'). "
    "Only modify the text AFTER the colon. Make it persuasive, punchy, and professional. "
    "IMPORTANT: Use a maximum of 1 em dash (—) across ALL the copy combined. Only one single em dash "
    "is allowed in the entire output. For everything else, use commas, periods, colons, or other punctuation. "
    "Return the result in the exact same structured format, one item per line."
)

SEO_TRANSLATOR_SYSTEM = (
    "You are an expert SEO translator specializing in {lang}. "
    "Translate the following SEO metadata from English to {lang}. "
    "For the URL slug, create a short, SEO-friendly slug in {lang} using only "
    "lowercase letters, numbers, and hyphens. "
    "CRITICAL: Keep the labels in English exactly as they appear "
    "(e.g. 'Page Title:', 'Meta Description:', 'URL Slug:'). "
    "Only translate the text AFTER the colon. Return in the same structured format."
)

SEO_COPYWRITER_SYSTEM = (
    "You are a world-class SEO copywriter and native {lang} speaker. "
    "Polish the following translated SEO metadata into compelling, natural {lang} copy "
    "optimized for search engines. The URL slug should be concise and keyword-rich in {lang}. "
    "CRITICAL: Keep the labels in English exactly as they appear "
    "(e.g. 'Page Title:', 'Meta Description:', 'URL Slug:'). "
    "Only modify the text AFTER the colon. Return in the same structured format."
)

ALT_TEXT_TRANSLATOR_SYSTEM = (
    "You are an expert translator specializing in {lang}. "
    "Translate the following image alt text from English to {lang}. "
    "Preserve descriptive accuracy for accessibility. "
    "CRITICAL: Keep the labels in English exactly as they appear (e.g. 'IMG_1:', 'IMG_2:'). "
    "Only translate the text AFTER the colon. Return in the same structured format."
)

ALT_TEXT_COPYWRITER_SYSTEM = (
    "You are a native {lang} copywriter specializing in accessible content. "
    "Polish these translated image alt texts to sound natural in {lang} "
    "while remaining descriptive and accessible. "
    "CRITICAL: Keep the labels in English exactly as they appear (e.g. 'IMG_1:', 'IMG_2:'). "
    "Only modify the text AFTER the colon. Return in the same structured format."
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def get_safe_path(base_dir: Path, name: str, lang: str, ext: str = ".xlsx") -> Path:
    """Return a file path that doesn't overwrite existing files."""
    path = base_dir / f"{name} - {lang}{ext}"
    if not path.exists():
        return path
    i = 1
    while True:
        path = base_dir / f"{name} - {lang} ({i}){ext}"
        if not path.exists():
            return path
        i += 1


def derive_page_name(url: str, soup: BeautifulSoup) -> str:
    """Derive a human-readable page name from the URL or H1."""
    h1 = soup.find("h1")
    if h1 and h1.get_text(strip=True):
        # Use first few words of H1, cleaned up
        text = h1.get_text(strip=True)
        # Truncate to something reasonable for a filename
        words = text.split()[:6]
        name = " ".join(words)
    else:
        # Fallback to URL slug
        path = urlparse(url).path.rstrip("/")
        name = path.split("/")[-1] if path else "homepage"
        name = name.replace("-", " ").title()
    # Clean for filesystem
    name = re.sub(r'[<>:"/\\|?*]', '', name)
    return name


# ---------------------------------------------------------------------------
# Scraping
# ---------------------------------------------------------------------------
def scrape_page(url: str) -> dict:
    """Scrape a Paperturn page and return structured content."""
    print(f"  Scraping {url} ...")
    resp = requests.get(url, timeout=30, headers={"User-Agent": "PaperturnTranslator/1.0"})
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "lxml")

    # --- Page metadata ---
    title_tag = soup.find("title")
    title = title_tag.get_text(strip=True) if title_tag else ""
    meta_desc_tag = soup.find("meta", attrs={"name": "description"})
    meta_desc = meta_desc_tag["content"].strip() if meta_desc_tag and meta_desc_tag.get("content") else ""

    # --- Main content (skip nav, header, footer) ---
    # Remove nav, header, footer elements
    for tag in soup.find_all(["nav", "header", "footer"]):
        tag.decompose()
    # Also remove common nav/footer classes
    for cls in ["navbar", "nav-wrapper", "footer", "site-footer", "site-header"]:
        for tag in soup.find_all(class_=re.compile(cls, re.I)):
            tag.decompose()

    # Extract content items
    content_items = []
    content_items.append(("Title", title))
    content_items.append(("Meta Description", meta_desc))

    # Find main content area
    main = soup.find("main") or soup.find(role="main") or soup.find("body")
    if not main:
        main = soup

    # Walk section_content divs in DOM order — these are Paperturn's CMS content blocks.
    # Each contains a mix of headings, paragraphs, bare text nodes, images, and links.
    from bs4 import NavigableString

    tag_map = {
        "h1": "H1", "h2": "H2", "h3": "H3", "h4": "H4", "h5": "H5", "h6": "H6",
        "p": "P",
    }
    cta_keywords = {"trial", "demo", "start", "book", "contact", "sign up", "get started", "free"}
    cta_hrefs = {"free-trial", "demo", "sign-up", "contact", "book-a-demo", "get-started"}
    # Trust signals are very specific phrases — match the full line, not substrings
    trust_phrases = {
        "dedicated in-trial support", "no credit card required", "cancel anytime",
        "trusted by", "organizations worldwide", "businesses worldwide",
        "no commitment", "money back guarantee", "14-day free trial",
    }

    seen_texts = set()

    def _is_trust_signal(text: str) -> bool:
        """Check if text is a known trust signal line."""
        lower = text.lower()
        return any(phrase in lower for phrase in trust_phrases)

    def _is_testimonial_attribution(text: str) -> bool:
        """Check if text looks like a name/job title attribution."""
        if len(text) < 5 or len(text) > 150:
            return False
        # Must be relatively short (job titles, not paragraphs)
        if len(text) > 80:
            return False
        # Use word-boundary matching to avoid false positives like "Leads"
        job_patterns = [
            r'\bmanager\b', r'\bdirector\b', r'\bpresident\b', r'\bceo\b',
            r'\bcfo\b', r'\bcoo\b', r'\bcoordinator\b', r'\bspecialist\b',
            r'\bofficer\b', r'\bhead of\b', r'\bvp\b', r'\bvice president\b',
            r'\bfounder\b', r'\beditor\b', r'\blead\b', r'\bchief\b',
        ]
        lower = text.lower()
        return any(re.search(pat, lower) for pat in job_patterns)

    def _add_item(tag_label: str, text: str):
        """Add a content item if not a duplicate."""
        if text and text not in seen_texts and len(text) > 1:
            seen_texts.add(text)
            content_items.append((tag_label, text))

    # Collect all image alt texts so we can skip them if they appear as P text
    all_alt_texts = {img.get("alt", "").strip() for img in main.find_all("img") if img.get("alt")}
    alt_text_phrases = ["animation showing", "animation demonstrating", "screenshot of",
                        "image of", "photo of", "illustration of", "icon of",
                        "duplicated for seamless loop", "clone of"]

    def _process_card(card_el):
        """Extract text from card-style links (feature-card, solution-card, etc.)."""
        for child in card_el.find_all(["h3", "h4", "h5"]):
            text = child.get_text(strip=True)
            if text:
                _add_item(tag_map.get(child.name, "H4"), text)
        for child in card_el.find_all("p"):
            text = child.get_text(strip=True)
            if text and len(text) >= 5 and text not in all_alt_texts:
                _add_item("P", text)

    def _process_showcase_card(card_el):
        """Extract flipbook showcase card title + author."""
        if card_el.get("aria-hidden") == "true":
            return
        title_el = card_el.find("p", class_="fb-showcase-card-title")
        sub_el = card_el.find("p", class_="fb-showcase-card-subtitle")
        if title_el:
            title = title_el.get_text(strip=True)
            subtitle = sub_el.get_text(strip=True) if sub_el else ""
            combined = f"{title} — {subtitle}" if subtitle else title
            _add_item("Showcase", combined)

    def _process_section(container):
        """Process a section_content div, extracting items in DOM order."""
        for child in container.children:
            # Bare text nodes (descriptions after H3s, attributions, trust signals)
            if isinstance(child, NavigableString):
                text = child.strip()
                if not text or len(text) < 5:
                    continue
                if text in seen_texts:
                    continue
                if _is_trust_signal(text):
                    _add_item("Trust Signal", text)
                elif _is_testimonial_attribution(text):
                    _add_item("Attribution", text)
                elif len(text) >= 15:
                    _add_item("P", text)
                continue

            if not hasattr(child, "name") or not child.name:
                continue

            text = child.get_text(strip=True)
            if not text or len(text) <= 1:
                continue

            # Headings and paragraphs
            if child.name in tag_map:
                tag_label = tag_map[child.name]
                if tag_label == "P" and len(text) < 5:
                    continue
                # Skip P tags that only contain CTA links
                if tag_label == "P":
                    child_links = child.find_all("a")
                    link_text = "".join(a.get_text(strip=True) for a in child_links)
                    if link_text and link_text == text:
                        continue
                _add_item(tag_label, text)

            # CTA buttons / links — also handle card-style links inline
            elif child.name in ("button", "a"):
                classes = " ".join(child.get("class", []))

                # Feature cards / solution cards
                if "feature-card" in classes or "solution-card" in classes:
                    _process_card(child)
                    continue

                # Flipbook showcase cards
                if "fb-showcase-card" in classes:
                    _process_showcase_card(child)
                    continue

                if len(text) >= 100:
                    continue
                href = child.get("href", "")
                is_cta = (
                    child.name == "button"
                    or "btn" in classes.lower()
                    or "cta" in classes.lower()
                    or any(kw in text.lower() for kw in cta_keywords)
                    or any(kw in href.lower() for kw in cta_hrefs)
                )
                if is_cta:
                    _add_item("CTA Button", text)

            # Font tags used as headings (CMS pattern: <font class="h2">)
            elif child.name == "font":
                classes = " ".join(child.get("class", []))
                handled = False
                for tag_name, label in [("h1", "H1"), ("h2", "H2"), ("h3", "H3"), ("h4", "H4")]:
                    if tag_name in classes:
                        _add_item(label, text)
                        handled = True
                        break
                if not handled:
                    _process_section(child)

            # Nested div or span — recurse for bare text inside
            elif child.name in ("div", "span", "small", "figcaption", "li", "blockquote"):
                _process_section(child)

    # Track which elements we've already processed to avoid duplicates
    processed_elements = set()

    def _mark_processed(el):
        """Mark an element and all its descendants as processed."""
        processed_elements.add(id(el))
        for desc in el.descendants:
            if hasattr(desc, 'name'):
                processed_elements.add(id(desc))

    # Single DOM-order walk: iterate all descendants of main
    for element in main.descendants:
        if id(element) in processed_elements:
            continue

        # Skip NavigableStrings at the top level (handled inside _process_section)
        if isinstance(element, NavigableString):
            continue

        if not hasattr(element, "name") or not element.name:
            continue

        classes = " ".join(element.get("class", []))

        # section_content divs — process their children
        if element.name == "div" and "section_content" in classes:
            _process_section(element)
            _mark_processed(element)

        # Feature cards
        elif element.name == "a" and "feature-card" in classes:
            _process_card(element)
            _mark_processed(element)

        # Solution cards
        elif element.name == "a" and "solution-card" in classes:
            _process_card(element)
            _mark_processed(element)

        # Flipbook showcase cards
        elif element.name == "a" and "fb-showcase-card" in classes:
            _process_showcase_card(element)
            _mark_processed(element)

        # Section description paragraphs outside section_content
        elif element.name == "p" and "section-description" in classes:
            text = element.get_text(strip=True)
            if text and len(text) >= 10:
                _add_item("P", text)
            _mark_processed(element)

    # Filter out alt text that leaked into content
    content_items_filtered = []
    for tag, text in content_items:
        if tag == "P" and text in all_alt_texts:
            continue
        if tag == "P" and any(phrase in text.lower() for phrase in alt_text_phrases):
            continue
        content_items_filtered.append((tag, text))
    content_items.clear()
    content_items.extend(content_items_filtered)

    # Fallback: if we found very little, do a broad pass (non-CMS pages)
    if len(content_items) < 5:
        for element in main.find_all(list(tag_map.keys())):
            text = element.get_text(strip=True)
            if not text or len(text) <= 1:
                continue
            tag_label = tag_map[element.name]
            if tag_label == "P" and len(text) < 5:
                continue
            if text in all_alt_texts:
                continue
            _add_item(tag_label, text)

        for btn in main.find_all(["button", "a"]):
            text = btn.get_text(strip=True)
            if not text or len(text) <= 1 or len(text) >= 100:
                continue
            classes = " ".join(btn.get("class", []))
            href = btn.get("href", "")
            is_cta = (
                btn.name == "button"
                or "btn" in classes.lower()
                or "cta" in classes.lower()
                or any(kw in text.lower() for kw in cta_keywords)
                or any(kw in href.lower() for kw in cta_hrefs)
            )
            if is_cta:
                _add_item("CTA Button", text)

    # --- Images ---
    images = []
    for img in main.find_all("img"):
        src = img.get("src", "")
        if src and not src.startswith("data:"):
            src = urljoin(url, src)
        alt = img.get("alt", "").strip()
        if alt:  # Only include images that have alt text
            images.append((src, alt))

    # --- URL slug ---
    parsed = urlparse(url)
    slug = parsed.path.rstrip("/")

    page_name = derive_page_name(url, soup if soup.find("h1") else BeautifulSoup(resp.text, "lxml"))

    return {
        "url": url,
        "page_name": page_name,
        "content": content_items,
        "images": images,
        "seo": {
            "title": title,
            "meta_description": meta_desc,
            "slug": slug,
        },
    }


def discover_urls(subpath: str = None, crawl_all: bool = False) -> list:
    """Discover page URLs from sitemap or by crawling."""
    urls = set()

    # Try sitemap first
    sitemap_url = f"{BASE_URL}/sitemap.xml"
    try:
        resp = requests.get(sitemap_url, timeout=15)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, "lxml-xml")
            for loc in soup.find_all("loc"):
                url = loc.get_text(strip=True)
                if subpath and subpath not in urlparse(url).path:
                    continue
                if not crawl_all and not subpath:
                    continue
                urls.add(url)
    except Exception as e:
        print(f"  Warning: Could not fetch sitemap: {e}")

    # If sitemap didn't yield results, crawl the base page
    if not urls:
        crawl_url = f"{BASE_URL}{subpath}" if subpath else BASE_URL
        try:
            resp = requests.get(crawl_url, timeout=15)
            soup = BeautifulSoup(resp.text, "lxml")
            for a in soup.find_all("a", href=True):
                href = a["href"]
                full = urljoin(BASE_URL, href)
                parsed = urlparse(full)
                if parsed.netloc and "paperturn.com" in parsed.netloc:
                    if subpath and subpath not in parsed.path:
                        continue
                    if parsed.path and parsed.path != "/":
                        urls.add(full)
        except Exception as e:
            print(f"  Warning: Could not crawl {crawl_url}: {e}")

    return sorted(urls)


# ---------------------------------------------------------------------------
# Claude API Translation Pipeline
# ---------------------------------------------------------------------------
def call_claude(client: anthropic.Anthropic, system: str, user_msg: str) -> str:
    """Make a single Claude API call and return the response text."""
    response = client.messages.create(
        model=MODEL,
        max_tokens=4096,
        system=system,
        messages=[{"role": "user", "content": user_msg}],
    )
    return response.content[0].text


def get_system_prompts(lang: str) -> dict:
    """Get all system prompts with the target language filled in."""
    return {
        "translator": TRANSLATOR_SYSTEM.format(lang=lang),
        "copywriter": COPYWRITER_SYSTEM.format(lang=lang),
        "seo_translator": SEO_TRANSLATOR_SYSTEM.format(lang=lang),
        "seo_copywriter": SEO_COPYWRITER_SYSTEM.format(lang=lang),
        "alt_translator": ALT_TEXT_TRANSLATOR_SYSTEM.format(lang=lang),
        "alt_copywriter": ALT_TEXT_COPYWRITER_SYSTEM.format(lang=lang),
    }


def format_content_for_api(items: list) -> str:
    """Format content items as 'TAG: text' lines for the API."""
    lines = []
    for tag, text in items:
        lines.append(f"{tag}: {text}")
    return "\n".join(lines)


def build_char_limit_instruction(content_items: list, char_limit_pct: int) -> str:
    """Build a character limit instruction string for the copywriter prompt."""
    lines = ["CHARACTER LIMITS: Each line must not exceed the max characters shown below."]
    for tag, text in content_items:
        max_chars = int(len(text) * (1 + char_limit_pct / 100))
        lines.append(f"  {tag}: max {max_chars} characters")
    lines.append("If the translation is too long, rephrase to fit. Do not truncate mid-sentence.")
    return "\n".join(lines)


def parse_api_response(response: str, expected_tags: list) -> list:
    """Parse 'TAG: text' response back into list of (tag, text) tuples."""
    results = []
    lines = response.strip().split("\n")

    # Sort tags by length descending so longer tags match first (e.g., "Page Title" before "P")
    sorted_tags = sorted(expected_tags, key=len, reverse=True)

    # Build a buffer for multi-line values
    current_tag = None
    current_text = []

    for line in lines:
        # Try to match a tag prefix
        matched = False
        for tag in sorted_tags:
            prefix = f"{tag}:"
            if line.startswith(prefix):
                # Save previous
                if current_tag is not None:
                    results.append((current_tag, " ".join(current_text).strip()))
                current_tag = tag
                current_text = [line[len(prefix):].strip()]
                matched = True
                break
        if not matched and current_tag is not None:
            current_text.append(line.strip())

    # Don't forget the last one
    if current_tag is not None:
        results.append((current_tag, " ".join(current_text).strip()))

    return results


SEO_REWRITE_SYSTEM = (
    "You are a world-class SEO copywriter and native {lang} speaker. "
    "You have been given marketing copy and a list of target SEO keywords. "
    "Your job is to naturally weave the most relevant keywords into the copy "
    "where they fit organically. Do NOT force keywords in — only use them where "
    "they sound natural. Prioritize higher-volume keywords. "
    "Do not change the meaning or tone of the copy. "
    "CRITICAL: Keep the labels/tags in English exactly as they appear. "
    "Only modify the text AFTER the colon. "
    "IMPORTANT: Use a maximum of 1 em dash (—) across ALL the copy combined. "
    "Return the result in the exact same structured format, one item per line."
)


def translate_content(client: anthropic.Anthropic, content_items: list,
                      translator_prompt: str, copywriter_prompt: str,
                      lang: str, char_limit_pct: int = None,
                      seo_keywords: list = None) -> tuple:
    """Two or three-step pipeline. Returns (translations, final_copies)."""
    if not content_items:
        return [], []

    expected_tags = [tag for tag, _ in content_items]
    formatted = format_content_for_api(content_items)

    # Step 1: Translate
    print(f"    Translating content ({len(content_items)} items) ...")
    translation_raw = call_claude(client, translator_prompt, formatted)
    translations = parse_api_response(translation_raw, expected_tags)

    # Step 2: Copywrite (fresh call)
    print(f"    Copywriting content ...")
    user_msg = translation_raw
    if char_limit_pct is not None:
        char_instructions = build_char_limit_instruction(content_items, char_limit_pct)
        user_msg = f"{translation_raw}\n\n{char_instructions}"
    final_raw = call_claude(client, copywriter_prompt, user_msg)
    final_copies = parse_api_response(final_raw, expected_tags)

    # Step 3: SEO keyword integration (fresh call, only if keywords provided)
    seo_copies = []
    if seo_keywords:
        print(f"    SEO rewrite with {len(seo_keywords)} keywords ...")
        kw_list = "\n".join(
            f"- {kw['keyword']} (volume: {kw['volume']})"
            for kw in seo_keywords
        )
        seo_prompt = SEO_REWRITE_SYSTEM.format(lang=lang)
        seo_user_msg = f"{final_raw}\n\nTARGET SEO KEYWORDS (prioritize higher volume):\n{kw_list}"
        if char_limit_pct is not None:
            char_instructions = build_char_limit_instruction(content_items, char_limit_pct)
            seo_user_msg = f"{seo_user_msg}\n\n{char_instructions}"
        seo_raw = call_claude(client, seo_prompt, seo_user_msg)
        seo_copies = parse_api_response(seo_raw, expected_tags)

    return translations, final_copies, seo_copies


def translate_images(client: anthropic.Anthropic, images: list,
                     translator_prompt: str, copywriter_prompt: str,
                     lang: str, char_limit_pct: int = None) -> tuple:
    """Translate image alt text through the two-step pipeline."""
    if not images:
        return [], []

    # Format as numbered items
    items = [(f"IMG_{i+1}", alt) for i, (_, alt) in enumerate(images)]
    expected_tags = [tag for tag, _ in items]
    formatted = format_content_for_api(items)

    print(f"    Translating {len(images)} image alt texts ...")
    translation_raw = call_claude(client, translator_prompt, formatted)
    translations = parse_api_response(translation_raw, expected_tags)

    print(f"    Copywriting image alt texts ...")
    user_msg = translation_raw
    if char_limit_pct is not None:
        char_instructions = build_char_limit_instruction(items, char_limit_pct)
        user_msg = f"{translation_raw}\n\n{char_instructions}"
    final_raw = call_claude(client, copywriter_prompt, user_msg)
    final_copies = parse_api_response(final_raw, expected_tags)

    return translations, final_copies


def translate_seo(client: anthropic.Anthropic, seo: dict,
                  translator_prompt: str, copywriter_prompt: str,
                  lang: str, char_limit_pct: int = None) -> tuple:
    """Translate SEO metadata through the two-step pipeline."""
    items = [
        ("Page Title", seo["title"]),
        ("Meta Description", seo["meta_description"]),
        ("URL Slug", seo["slug"]),
    ]
    expected_tags = [tag for tag, _ in items]
    formatted = format_content_for_api(items)

    print(f"    Translating SEO metadata ...")
    translation_raw = call_claude(client, translator_prompt, formatted)
    translations = parse_api_response(translation_raw, expected_tags)

    print(f"    Copywriting SEO metadata ...")
    user_msg = translation_raw
    if char_limit_pct is not None:
        char_instructions = build_char_limit_instruction(items, char_limit_pct)
        user_msg = f"{translation_raw}\n\n{char_instructions}"
    final_raw = call_claude(client, copywriter_prompt, user_msg)
    final_copies = parse_api_response(final_raw, expected_tags)

    return translations, final_copies


# ---------------------------------------------------------------------------
# SEMrush Integration (optional)
# ---------------------------------------------------------------------------
def extract_seed_keywords(client: anthropic.Anthropic, content_items: list, lang: str) -> list:
    """Use Claude to extract seed keywords from page content, translated into target language."""
    # Build a summary of the page content
    page_text = "\n".join(f"{tag}: {text}" for tag, text in content_items[:15])

    response = call_claude(
        client,
        f"You are an SEO keyword research expert fluent in {lang}.",
        f"Based on this page content, generate 8 short seed keywords in {lang} (1-2 words each) "
        f"that someone searching for this type of product/service would use. "
        f"Keep them broad enough to have search volume. "
        f"Return ONLY the keywords, one per line, no numbers or bullets. "
        f"Keywords must be in {lang}.\n\n{page_text}",
    )
    keywords = [line.strip() for line in response.strip().split("\n") if line.strip()]
    return keywords[:8]


def fetch_semrush_keywords(api_key: str, url: str, lang: str,
                           client: anthropic.Anthropic = None,
                           content_items: list = None) -> list:
    """Fetch keyword recommendations from SEMrush API using seed keywords from the page."""
    lang_db_map = {
        "spanish": "es", "french": "fr", "german": "de",
        "danish": "dk", "swedish": "se", "italian": "it",
        "portuguese": "br", "dutch": "nl",
    }
    db = lang_db_map.get(lang.lower(), "us")

    # Step 1: Extract seed keywords using Claude
    seed_keywords = []
    if client and content_items:
        print(f"    Extracting seed keywords for {lang} ...")
        seed_keywords = extract_seed_keywords(client, content_items, lang)
        print(f"    Seed keywords: {seed_keywords}")

    if not seed_keywords:
        # Fallback: use URL path words
        path = urlparse(url).path.replace("/", " ").replace("-", " ").strip()
        seed_keywords = [path] if path else ["flipbook"]

    # Step 2: Query SEMrush for each seed keyword
    all_keywords = {}
    for seed in seed_keywords:
        for query_type in ["phrase_related", "phrase_this"]:
            try:
                api_url = (
                    f"https://api.semrush.com/"
                    f"?type={query_type}"
                    f"&key={api_key}"
                    f"&phrase={requests.utils.quote(seed)}"
                    f"&database={db}"
                    f"&display_limit=10"
                    f"&export_columns=Ph,Nq,Cp,Co"
                )
                resp = requests.get(api_url, timeout=30)
                if resp.status_code == 200 and not resp.text.strip().startswith("ERROR"):
                    reader = csv.reader(resp.text.strip().split("\n"), delimiter=";")
                    next(reader, None)  # Skip header
                    for row in reader:
                        if len(row) >= 2:
                            kw = row[0]
                            vol = row[1] if len(row) > 1 else "0"
                            cpc = row[2] if len(row) > 2 else "0"
                            comp = row[3] if len(row) > 3 else "0"
                            # Deduplicate, keep highest volume
                            if kw not in all_keywords or int(vol) > int(all_keywords[kw].get("volume", "0") or "0"):
                                all_keywords[kw] = {
                                    "keyword": kw,
                                    "volume": vol,
                                    "cpc": cpc,
                                    "competition": comp,
                                    "seed": seed,
                                }
            except Exception as e:
                print(f"    SEMrush error for '{seed}' ({query_type}): {e}")

    # Sort by search volume descending
    results = sorted(all_keywords.values(), key=lambda x: int(x.get("volume", "0") or "0"), reverse=True)
    print(f"    SEMrush raw total: {len(results)} unique keywords")

    # Filter with Claude to keep only relevant keywords
    if client and content_items and results:
        top_candidates = results[:50]
        kw_list = "\n".join(f"- {kw['keyword']} (vol: {kw['volume']})" for kw in top_candidates)
        page_summary = " ".join(text for _, text in content_items[:5])

        filter_response = call_claude(
            client,
            "You are an SEO expert. Filter keyword lists for relevance.",
            f"This page is about: {page_summary[:300]}\n\n"
            f"From this keyword list, return ONLY the keywords that are directly relevant "
            f"to digital catalogs, flipbooks, PDF publishing, product catalogs, or industrial/manufacturing "
            f"content management. Remove brand names, locations, and unrelated terms. "
            f"Return one keyword per line, nothing else.\n\n{kw_list}",
        )
        relevant = {line.strip().lstrip("- ").lower() for line in filter_response.strip().split("\n") if line.strip()}
        filtered = [kw for kw in top_candidates if kw["keyword"].lower() in relevant]
        print(f"    SEMrush filtered: {len(filtered)} relevant keywords")
        return filtered[:30]

    return results[:30]


# ---------------------------------------------------------------------------
# Excel Output
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Calibri", bold=True, size=11)
HEADER_FILL = PatternFill(start_color="117681", end_color="117681", fill_type="solid")
HEADER_FONT_WHITE = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header(ws, num_cols):
    """Apply header styling to the first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = WRAP


def write_xlsx(output_path: Path, page_data: dict, content_translations: tuple,
               image_translations: tuple, seo_translations: tuple,
               semrush_keywords: list = None):
    """Write all translation data to an Excel workbook."""
    wb = Workbook()

    # --- Sheet 1: Content ---
    ws_content = wb.active
    ws_content.title = "Content"
    content_items = page_data["content"]
    translations, final_copies, seo_copies = content_translations

    has_seo_col = bool(seo_copies)
    if has_seo_col:
        ws_content.append(["Text Tag", "English", "Translation", "Final Copy", "SEO Copy"])
        style_header(ws_content, 5)
    else:
        ws_content.append(["Text Tag", "English", "Translation", "Final Copy"])
        style_header(ws_content, 4)

    for i, (tag, english) in enumerate(content_items):
        trans_text = translations[i][1] if i < len(translations) else ""
        final_text = final_copies[i][1] if i < len(final_copies) else ""
        row = [tag, english, trans_text, final_text]
        if has_seo_col:
            seo_text = seo_copies[i][1] if i < len(seo_copies) else ""
            row.append(seo_text)
        ws_content.append(row)

    # Auto-width columns
    for col_letter in ["A", "B", "C", "D", "E"]:
        if col_letter in ws_content.column_dimensions:
            ws_content.column_dimensions[col_letter].width = 40
    ws_content.column_dimensions["A"].width = 18

    # --- Sheet 2: Images ---
    ws_images = wb.create_sheet("Images")
    ws_images.append(["Image URL", "English Alt Text", "Translation", "Final Copy"])
    style_header(ws_images, 4)

    images = page_data["images"]
    img_translations, img_final = image_translations

    for i, (src, alt) in enumerate(images):
        trans_text = img_translations[i][1] if i < len(img_translations) else ""
        final_text = img_final[i][1] if i < len(img_final) else ""
        ws_images.append([src, alt, trans_text, final_text])

    for col_letter in ["A", "B", "C", "D"]:
        ws_images.column_dimensions[col_letter].width = 40

    # --- Sheet 3: SEO ---
    ws_seo = wb.create_sheet("SEO")
    ws_seo.append(["Field", "English", "Translation", "Final Copy"])
    style_header(ws_seo, 4)

    seo = page_data["seo"]
    seo_items = [
        ("Page Title", seo["title"]),
        ("Meta Description", seo["meta_description"]),
        ("URL Slug", seo["slug"]),
    ]
    seo_trans, seo_final = seo_translations

    for i, (field, english) in enumerate(seo_items):
        trans_text = seo_trans[i][1] if i < len(seo_trans) else ""
        final_text = seo_final[i][1] if i < len(seo_final) else ""
        ws_seo.append([field, english, trans_text, final_text])

    # SEMrush keywords section
    if semrush_keywords:
        ws_seo.append([])  # blank row
        ws_seo.append(["SEMrush Keywords", "Volume", "CPC", "Competition"])
        # Style the keyword header row
        kw_header_row = ws_seo.max_row
        for col in range(1, 5):
            cell = ws_seo.cell(row=kw_header_row, column=col)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
        for kw in semrush_keywords:
            ws_seo.append([
                kw.get("keyword", ""),
                kw.get("volume", ""),
                kw.get("cpc", ""),
                kw.get("competition", ""),
            ])
    else:
        ws_seo.append(["SEMrush Keywords", "—", "—", "—"])

    for col_letter in ["A", "B", "C", "D"]:
        ws_seo.column_dimensions[col_letter].width = 45
    ws_seo.column_dimensions["A"].width = 30

    # Save
    wb.save(output_path)
    print(f"  Saved: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Batch translate Paperturn pages into target languages.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python translate.py --lang Spanish --urls urls.txt
  python translate.py --lang French --subpath /industries/
  python translate.py --lang German --all
        """,
    )
    parser.add_argument("--lang", required=True, help="Target language (e.g., Spanish, French)")
    parser.add_argument("--urls", help="Path to text file with one URL per line")
    parser.add_argument("--subpath", help="Crawl pages under a subpath (e.g., /industries/)")
    parser.add_argument("--all", action="store_true", help="Crawl entire site")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output directory")
    parser.add_argument("--semrush", action="store_true", help="Enable SEMrush keyword analysis (prompts per URL)")
    parser.add_argument("--char-limit", type=int, default=None, metavar="PCT",
                        help="Enforce character limit on final copy (e.g., 10 = allow 10%% overshoot vs English)")

    args = parser.parse_args()

    # Validate input mode
    if not args.urls and not args.subpath and not args.all:
        parser.error("Provide one of: --urls <file>, --subpath <path>, or --all")

    # Check API key
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("Error: ANTHROPIC_API_KEY environment variable not set.")
        print("  export ANTHROPIC_API_KEY='your-key-here'")
        sys.exit(1)

    semrush_key = os.environ.get("SEMRUSH_API_KEY") if args.semrush else None
    if args.semrush and not semrush_key:
        print("Error: SEMRUSH_API_KEY environment variable not set.")
        sys.exit(1)

    # Gather URLs
    urls = []
    if args.urls:
        url_file = Path(args.urls)
        if not url_file.exists():
            print(f"Error: URL file not found: {args.urls}")
            sys.exit(1)
        urls = [line.strip() for line in url_file.read_text().splitlines() if line.strip() and not line.startswith("#")]
    elif args.subpath or args.all:
        print(f"Discovering URLs {'for ' + args.subpath if args.subpath else '(full site)'}...")
        urls = discover_urls(subpath=args.subpath, crawl_all=args.all)

    if not urls:
        print("No URLs found to process.")
        sys.exit(1)

    print(f"\nFound {len(urls)} page(s) to translate into {args.lang}:\n")
    for u in urls:
        print(f"  - {u}")
    print()

    # Setup
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)
    client = anthropic.Anthropic(api_key=api_key)

    # Process each page
    for i, url in enumerate(urls, 1):
        print(f"\n[{i}/{len(urls)}] Processing: {url}")

        # 1. Scrape
        try:
            page_data = scrape_page(url)
        except Exception as e:
            print(f"  Error scraping {url}: {e}")
            continue

        # 2. Get system prompts for this page (fresh per page)
        prompts = get_system_prompts(args.lang)

        # 3. SEMrush (run first so keywords feed into content rewrite)
        semrush_keywords = []
        if args.semrush and semrush_key:
            confirm = input(f"\n  Run SEMrush analysis for {url}? [y/N]: ").strip().lower()
            if confirm == "y":
                print(f"    Fetching SEMrush keywords ...")
                semrush_keywords = fetch_semrush_keywords(
                    semrush_key, url, args.lang,
                    client=client, content_items=page_data["content"],
                )
                print(f"    Found {len(semrush_keywords)} keywords")

        # 4. Content translation + copywriting + SEO rewrite
        try:
            content_translations = translate_content(
                client, page_data["content"],
                prompts["translator"],
                prompts["copywriter"],
                args.lang,
                char_limit_pct=args.char_limit,
                seo_keywords=semrush_keywords if semrush_keywords else None,
            )
        except Exception as e:
            print(f"  Error translating content: {e}")
            content_translations = ([], [], [])

        # 5. Image alt text translation + copywriting
        try:
            image_translations = translate_images(
                client, page_data["images"],
                prompts["alt_translator"],
                prompts["alt_copywriter"],
                args.lang,
                char_limit_pct=args.char_limit,
            )
        except Exception as e:
            print(f"  Error translating images: {e}")
            image_translations = ([], [])

        # 6. SEO translation + copywriting
        try:
            seo_translations = translate_seo(
                client, page_data["seo"],
                prompts["seo_translator"],
                prompts["seo_copywriter"],
                args.lang,
                char_limit_pct=args.char_limit,
            )
        except Exception as e:
            print(f"  Error translating SEO: {e}")
            seo_translations = ([], [])

        # 7. Write Excel
        file_path = get_safe_path(output_dir, page_data["page_name"], args.lang)
        try:
            write_xlsx(file_path, page_data, content_translations,
                       image_translations, seo_translations, semrush_keywords)
        except Exception as e:
            print(f"  Error writing Excel: {e}")

    print(f"\nDone! {len(urls)} page(s) processed. Output: {output_dir}")


if __name__ == "__main__":
    main()
